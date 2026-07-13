from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from repositories.customer_repository import CustomerRepository
from repositories.product_repository import ProductRepository
from repositories.order_repository import OrderRepository


class ExcelService:

    def __init__(self):

        self.customer_repository = CustomerRepository()

        self.product_repository = ProductRepository()

        self.order_repository = OrderRepository()

    def format_sheet(
            self,
            ws):

        fill = PatternFill(

            fill_type="solid",

            start_color="0D6EFD"

        )

        font = Font(

            color="FFFFFF",

            bold=True

        )

        center = Alignment(

            horizontal="center"

        )

        for cell in ws[1]:

            cell.fill = fill

            cell.font = font

            cell.alignment = center

    def export_customers(self):

        workbook = Workbook()

        ws = workbook.active

        ws.title = "Customers"

        ws.append([

            "Customer ID",

            "Name",

            "Phone",

            "Email",

            "Balance"

        ])

        customers = self.customer_repository.get_all_customers()

        for customer in customers:

            ws.append([

                customer.customer_id,

                customer.name,

                customer.phone,

                customer.email,

                customer.balance

            ])

        self.format_sheet(ws)

        workbook.save("exports/customers.xlsx")

    def export_products(self):

        workbook = Workbook()

        ws = workbook.active

        ws.title = "Products"

        ws.append([

            "Product ID",

            "Name",

            "Price",

            "Stock"

        ])

        products = self.product_repository.get_all_products()

        for product in products:

            ws.append([

                product.product_id,

                product.name,

                product.price,

                product.stock

            ])

        self.format_sheet(ws)

        workbook.save("exports/products.xlsx")

    def export_orders(self):

        workbook = Workbook()

        ws = workbook.active

        ws.title = "Orders"

        ws.append([

            "Order ID",

            "Customer",

            "Product",

            "Quantity",

            "Total Price"

        ])

        orders = self.order_repository.get_all_orders()

        for order in orders:

            ws.append([

                order.order_id,

                order.customer.name,

                order.product.name,

                order.quantity,

                order.total_price

            ])

        self.format_sheet(ws)

        workbook.save("exports/orders.xlsx")